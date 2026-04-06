from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any
import re
from pathlib import Path

# Illegal characters for Excel/XML
# See: https://stackoverflow.com/questions/13010323/illegal-characters-in-openpyxl-excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010\013\014\016-\037]')

def clean_for_excel(val):
    """Strip illegal characters that break openpyxl/XML"""
    if not isinstance(val, str):
        return val
    # Some bank statements contain characters like \0 (null) or other control chars
    return ILLEGAL_CHARACTERS_RE.sub('', val)

def generate_excel(statement_results: List[Dict[str, Any]], combined_validation: Dict[str, Any], output_path: Path):
    """
    Generate Excel file with multiple worksheets for each statement
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    for idx, stmt in enumerate(statement_results):
        transactions = stmt.get("transactions", [])
        validation = stmt.get("validation", {})
        metadata = stmt.get("metadata", {})
        
        # Sheet name: Account No or Index
        acc_no = metadata.get("account_no") or f"Stmt {idx+1}"
        ws = wb.create_sheet(title=str(acc_no)[:31]) # Max 31 chars
        
        # Headers
        headers = ['Date', 'Value Date', 'Reference', 'Description', 'Category', 'Debit', 'Credit', 'Balance']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add transactions
        for txn in transactions:
            desc = (
                txn.get('description') or
                txn.get('remarks') or
                txn.get('originating_branch') or ''
            )
            ws.append([
                clean_for_excel(str(txn.get('date', ''))),
                clean_for_excel(str(txn.get('value_date', ''))),
                clean_for_excel(txn.get('reference', '')),
                clean_for_excel(desc),
                clean_for_excel(txn.get('category', 'Unallocated')),
                txn.get('debit', 0),
                txn.get('credit', 0),
                txn.get('balance', 0)
            ])
        
        # Add summary section
        summary_row = len(transactions) + 3
        ws[f'A{summary_row}'] = 'STATEMENT SUMMARY'
        ws[f'A{summary_row}'].font = Font(bold=True, size=14)
        
        ws[f'A{summary_row + 1}'] = 'Account Name:'
        ws[f'B{summary_row + 1}'] = metadata.get("account_name", "N/A")
        
        ws[f'A{summary_row + 2}'] = 'Total Debit:'
        ws[f'B{summary_row + 2}'] = validation.get('extracted_total_debit', 0)
        ws[f'B{summary_row + 2}'].number_format = '#,##0.00'
        
        ws[f'A{summary_row + 3}'] = 'Total Credit:'
        ws[f'B{summary_row + 3}'] = validation.get('extracted_total_credit', 0)
        ws[f'B{summary_row + 3}'].number_format = '#,##0.00'
        
        ws[f'A{summary_row + 4}'] = 'Validation Status:'
        ws[f'B{summary_row + 4}'] = metadata.get('validation_status', validation.get('status', 'Unknown'))
        
        # If there are gaps found during audit, add an Audit Sheet
        mismatch_details = metadata.get("mismatch_details")
        if mismatch_details:
            audit_ws = wb.create_sheet(title=f"Audit - {str(acc_no)[:20]}")
            audit_ws.append(['Impacted Page', 'Date', 'Description', 'Expected Balance', 'Extracted Balance', 'Difference'])
            for gap in mismatch_details:
                audit_ws.append([
                    gap.get('page', 'Unknown'),
                    gap.get('date', 'Unknown'),
                    gap.get('description', 'Unknown'),
                    gap.get('expected', 0),
                    gap.get('claimed', 0),
                    gap.get('diff', 0)
                ])
            # Color the difference red
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for row in range(2, len(mismatch_details) + 2):
                audit_ws[f'F{row}'].fill = red_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
    
    # Save
    wb.save(output_path)
