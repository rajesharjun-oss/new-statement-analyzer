import re
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Helpers
# -----------------------------

MONEY_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d{2})$|^-?\d+(?:\.\d{2})$")
DATE_RE = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")  # e.g. 05-Jun-2025


def parse_money(s: str) -> Optional[float]:
    s = (s or "").strip()
    if not s:
        return None
    s = s.replace(" ", "")
    if not MONEY_RE.match(s):
        return None
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None


def looks_like_date(s: str) -> bool:
    return bool(DATE_RE.match((s or "").strip()))


def safe_join(parts: List[str]) -> str:
    txt = " ".join([p for p in (p.strip() for p in parts) if p])
    return re.sub(r"\s+", " ", txt).strip()


def normalize_ref(desc: str) -> Tuple[str, str]:
    """
    Pull out "REFNO:...." (or similar) from description if present.
    Returns (reference, cleaned_description)
    """
    if not desc:
        return "", ""
    # Common pattern in your screenshots: "REFNO:A01ECTS2515300007 ..."
    m = re.search(r"\bREF(?:NO)?[:\s]*([A-Za-z0-9]+)\b", desc, flags=re.IGNORECASE)
    if not m:
        return "", desc.strip()
    ref = m.group(1).strip()
    cleaned = re.sub(r"\bREF(?:NO)?[:\s]*" + re.escape(ref) + r"\b", "", desc, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,|-")
    return ref, cleaned


@dataclass
class ColumnBounds:
    # x-intervals used to assign words to columns
    txn_date: Tuple[float, float]
    description: Tuple[float, float]
    value_date: Tuple[float, float]
    debit: Tuple[float, float]
    credit: Tuple[float, float]
    balance: Tuple[float, float]


def midpoint(x0: float, x1: float) -> float:
    return (x0 + x1) / 2.0


def in_interval(x: float, interval: Tuple[float, float]) -> bool:
    return interval[0] <= x < interval[1]


# -----------------------------
# Core extraction
# -----------------------------

def _find_header_and_bounds(words: List[Dict]) -> Optional[ColumnBounds]:
    """
    Detect the table header on a page and infer x-bounds for each column.
    Strategy:
      - locate header words: "Transaction", "Date", "Description", "Value", "Debit", "Credit", "Balance"
      - use their x positions to create boundaries
    """
    # Build a simple line index by y (top)
    # pdfplumber words include: x0, x1, top, bottom, text
    # We'll find candidates near each other in y.
    if not words:
        return None

    # group words by "row" using y clustering
    words_sorted = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))
    rows: List[List[Dict]] = []
    row: List[Dict] = []
    last_top = None
    for w in words_sorted:
        t = w["top"]
        if last_top is None or abs(t - last_top) <= 2.0:
            row.append(w)
        else:
            rows.append(row)
            row = [w]
        last_top = t
    if row:
        rows.append(row)

    def row_text(r: List[Dict]) -> str:
        return " ".join([w["text"] for w in r])

    header_row = None
    for r in rows:
        t = row_text(r).lower()
        # Ecobank screenshot header looks like:
        # "Transaction Date  Description  Value Date  Debit  Credit  Balance"
        if ("transaction" in t and "date" in t and "description" in t and "value" in t
                and "debit" in t and "credit" in t and "balance" in t):
            header_row = r
            break

    if not header_row:
        return None

    # get approximate x positions of key labels in header row
    # We'll pick left edge x0 for each label group.
    def find_x(label_tokens: List[str]) -> Optional[float]:
        # find first occurrence in the row
        txts = [w["text"] for w in header_row]
        low = [t.lower() for t in txts]
        for i in range(len(low)):
            if low[i] == label_tokens[0]:
                # ensure sequence matches
                ok = True
                for j, tok in enumerate(label_tokens):
                    if i + j >= len(low) or low[i + j] != tok:
                        ok = False
                        break
                if ok:
                    return header_row[i]["x0"]
        return None

    x_txn = find_x(["transaction"])  # header often splits "Transaction" and "Date"
    x_desc = find_x(["description"])
    x_val = find_x(["value"])
    x_deb = find_x(["debit"])
    x_cre = find_x(["credit"])
    x_bal = find_x(["balance"])

    # Fallback: sometimes "Transaction Date" is two words and "Value Date" is two words
    if x_txn is None:
        x_txn = find_x(["transaction"])
    if x_val is None:
        x_val = find_x(["value"])

    if None in (x_txn, x_desc, x_val, x_deb, x_cre, x_bal):
        return None

    # Create boundaries based on the x positions, sorted
    # txn_date starts at 0 up to description start
    # description up to value start, etc.
    xs = sorted([x_txn, x_desc, x_val, x_deb, x_cre, x_bal])
    # We want boundaries in the natural order: txn < desc < val < debit < credit < balance
    # If order is weird (rare), bail out.
    if not (x_txn < x_desc < x_val < x_deb < x_cre < x_bal):
        return None

    # Add some padding to reduce boundary mis-assignments
    pad = 2.0
    return ColumnBounds(
        txn_date=(0, x_desc - pad),
        description=(x_desc - pad, x_val - pad),
        value_date=(x_val - pad, x_deb - pad),
        debit=(x_deb - pad, x_cre - pad),
        credit=(x_cre - pad, x_bal - pad),
        balance=(x_bal - pad, 10_000),
    )


def _extract_table_words(page) -> List[Dict]:
    # Keep it simple: pull words with decent tolerances
    return page.extract_words(
        x_tolerance=1.5,
        y_tolerance=2.0,
        keep_blank_chars=False,
        use_text_flow=True,
    )


def _words_to_rows(words: List[Dict], bounds: ColumnBounds) -> List[Dict]:
    """
    Convert page words into row dicts with raw strings for each column.
    Handles multi-line descriptions by emitting "continuation rows" (empty txn_date),
    which we later merge.
    """
    if not words:
        return []

    # Group words by y (line)
    words_sorted = sorted(words, key=lambda w: (round(w["top"], 1), w["x0"]))
    lines: List[List[Dict]] = []
    line: List[Dict] = []
    last_top = None
    for w in words_sorted:
        t = w["top"]
        if last_top is None or abs(t - last_top) <= 2.0:
            line.append(w)
        else:
            lines.append(line)
            line = [w]
        last_top = t
    if line:
        lines.append(line)

    rows: List[Dict] = []
    for ln in lines:
        # assign words to buckets by x midpoints
        txn_parts, desc_parts, val_parts, deb_parts, cre_parts, bal_parts = [], [], [], [], [], []
        for w in ln:
            x = midpoint(w["x0"], w["x1"])
            txt = w["text"].strip()
            if not txt:
                continue

            if in_interval(x, bounds.txn_date):
                txn_parts.append(txt)
            elif in_interval(x, bounds.description):
                desc_parts.append(txt)
            elif in_interval(x, bounds.value_date):
                val_parts.append(txt)
            elif in_interval(x, bounds.debit):
                deb_parts.append(txt)
            elif in_interval(x, bounds.credit):
                cre_parts.append(txt)
            elif in_interval(x, bounds.balance):
                bal_parts.append(txt)

        # ignore header line itself and other noise
        txn_txt = safe_join(txn_parts)
        desc_txt = safe_join(desc_parts)
        val_txt = safe_join(val_parts)
        deb_txt = safe_join(deb_parts)
        cre_txt = safe_join(cre_parts)
        bal_txt = safe_join(bal_parts)

        # Heuristic: keep only lines that look like table content:
        # - either has a txn date OR has money/balance columns OR is a continuation description line
        has_money = any(parse_money(x) is not None for x in [deb_txt, cre_txt, bal_txt])
        if looks_like_date(txn_txt) or has_money or desc_txt:
            # Critical guard: Value Date must be a date, not "50.00" etc.
            # If value_date looks like money and debit is empty, it's probably shifted;
            # BUT with bounds-based assignment this should be rare. Still guard hard.
            if val_txt and (parse_money(val_txt) is not None) and (not looks_like_date(val_txt)):
                # drop it rather than polluting Value Date
                val_txt = ""

            rows.append({
                "Txn Date": txn_txt,
                "Value Date": val_txt,
                "Description": desc_txt,
                "Debit_raw": deb_txt,
                "Credit_raw": cre_txt,
                "Balance_raw": bal_txt,
            })

    return rows


def _merge_continuations(rows: List[Dict]) -> List[Dict]:
    """
    Merge lines where Txn Date is empty into the previous transaction's Description.
    Also fixes cases where Description continues while amounts appear only on first line.
    """
    merged: List[Dict] = []
    current = None

    for r in rows:
        txn_date = (r.get("Txn Date") or "").strip()
        desc = (r.get("Description") or "").strip()

        # skip obvious non-transaction lines
        if desc.upper().startswith("OPENING BALANCE"):
            # treat as a real row if it has a txn date; otherwise ignore
            if looks_like_date(txn_date):
                current = r.copy()
                merged.append(current)
            continue

        if looks_like_date(txn_date):
            current = r.copy()
            merged.append(current)
        else:
            # continuation line
            if current is None:
                continue
            if desc:
                current["Description"] = safe_join([current.get("Description", ""), desc])

            # Sometimes continuation lines accidentally capture value date; keep first non-empty.
            if not current.get("Value Date") and r.get("Value Date"):
                if looks_like_date(r["Value Date"]):
                    current["Value Date"] = r["Value Date"]

            # Rare: amounts captured in continuation; fill if missing
            for k in ["Debit_raw", "Credit_raw", "Balance_raw"]:
                if (not (current.get(k) or "").strip()) and (r.get(k) or "").strip():
                    current[k] = r[k]

    return merged


def _finalize_rows(rows: List[Dict]) -> pd.DataFrame:
    out = []
    for r in rows:
        txn_date = (r.get("Txn Date") or "").strip()
        val_date = (r.get("Value Date") or "").strip()
        desc = (r.get("Description") or "").strip()

        debit = parse_money(r.get("Debit_raw", ""))
        credit = parse_money(r.get("Credit_raw", ""))
        balance = parse_money(r.get("Balance_raw", ""))

        # Only keep rows that look like real transactions
        if not looks_like_date(txn_date):
            continue
        if not desc and debit is None and credit is None and balance is None:
            continue

        ref, cleaned_desc = normalize_ref(desc)

        out.append({
            "Txn Date": txn_date,
            "Value Date": val_date if looks_like_date(val_date) else "",
            "Reference": ref,
            "Description": cleaned_desc,
            "Debit": float(debit) if debit is not None else 0.0,
            "Credit": float(credit) if credit is not None else 0.0,
            "Balance": float(balance) if balance is not None else 0.0,
        })

    df = pd.DataFrame(out)

    # Keep original order; also remove any exact duplicates that can appear at page boundaries
    if not df.empty:
        df = df.drop_duplicates(subset=["Txn Date", "Value Date", "Reference", "Description", "Debit", "Credit", "Balance"])
    return df


# -----------------------------
# Public API
# -----------------------------

def extract_ecobank_transactions(pdf_path: str) -> pd.DataFrame:
    """
    Extract Ecobank statement transactions into a normalized DataFrame.

    Output columns:
      Txn Date | Value Date | Reference | Description | Debit | Credit | Balance
    """
    all_rows: List[Dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            words = _extract_table_words(page)
            bounds = _find_header_and_bounds(words)
            if bounds is None:
                # Not a transaction table page (or header not found). Skip safely.
                continue

            # Keep only words below the header line to avoid mixing header into rows.
            # We'll find the approximate header Y by locating the word "Transaction" used in header.
            header_candidates = [w for w in words if w["text"].strip().lower() == "transaction"]
            header_top = min([w["top"] for w in header_candidates], default=None)
            if header_top is not None:
                words = [w for w in words if w["top"] > header_top + 4.0]

            page_rows = _words_to_rows(words, bounds)

            # Remove footer noise like "Page 2 | 10"
            cleaned = []
            for r in page_rows:
                d = (r.get("Description") or "").strip().lower()
                if d.startswith("page ") or d.endswith("| 10") or d.endswith("|10"):
                    continue
                cleaned.append(r)

            all_rows.extend(cleaned)

    merged = _merge_continuations(all_rows)
    df = _finalize_rows(merged)
    return df


def export_transactions_to_excel(df: pd.DataFrame, out_xlsx: str) -> None:
    """
    Write transactions to Excel with solid formatting.
    """
    if df is None or df.empty:
        # still create a file with headers
        df = pd.DataFrame(columns=["Txn Date", "Value Date", "Reference", "Description", "Debit", "Credit", "Balance"])

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Transactions")

    wb = load_workbook(out_xlsx)
    ws = wb["Transactions"]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths + formats
    col_map = {ws.cell(1, i).value: i for i in range(1, ws.max_column + 1)}

    def set_width(name: str, width: int):
        if name in col_map:
            ws.column_dimensions[get_column_letter(col_map[name])].width = width

    set_width("Txn Date", 14)
    set_width("Value Date", 14)
    set_width("Reference", 20)
    set_width("Description", 60)
    set_width("Debit", 16)
    set_width("Credit", 16)
    set_width("Balance", 18)

    money_fmt = '#,##0.00'
    for name in ["Debit", "Credit", "Balance"]:
        if name in col_map:
            idx = col_map[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, idx).number_format = money_fmt

    # Wrap description
    if "Description" in col_map:
        idx = col_map["Description"]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, idx).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_xlsx)


def extract_ecobank_to_excel(pdf_path: str, out_xlsx: str) -> pd.DataFrame:
    """
    One-shot convenience function.
    """
    df = extract_ecobank_transactions(pdf_path)
    export_transactions_to_excel(df, out_xlsx)
    return df


# -----------------------------
# Example CLI usage
# -----------------------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        pdf_path = "ecobank_statement.pdf"
    
    out_xlsx = pdf_path.replace(".pdf", ".xlsx")
    if out_xlsx == pdf_path:
        out_xlsx += ".xlsx"
        
    try:
        df = extract_ecobank_to_excel(pdf_path, out_xlsx)
        print(f"✅ Successfully extracted {len(df)} transactions -> {out_xlsx}")
    except Exception as e:
        print(f"❌ Error: {e}")
